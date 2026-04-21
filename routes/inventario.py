import os, pandas as pd
from flask import Blueprint, render_template, request, redirect, url_for, flash, session, jsonify, current_app
from sqlalchemy import func
from datetime import datetime
from werkzeug.utils import secure_filename
from extensions import db
from models import Repuesto, Venta, HistorialStock

inventario_bp = Blueprint('inventario', __name__)
ROLES_PRIVILEGIADOS = ('admin', 'supervisor')
ALLOWED_EXT = {'png','jpg','jpeg','webp','gif'}

def puede_ver_costos():
    return session.get('user_rol') in ROLES_PRIVILEGIADOS

def allowed_file(filename):
    return '.' in filename and filename.rsplit('.',1)[1].lower() in ALLOWED_EXT

def _historial(repuesto_id, ant, nuevo, accion):
    uid = session.get('user_id')
    db.session.add(HistorialStock(repuesto_id=repuesto_id, usuario_id=uid,
                                  stock_anterior=ant, stock_nuevo=nuevo, accion=accion))

def _stats():
    hoy        = datetime.utcnow().date()
    ventas_hoy = Venta.query.filter(func.date(Venta.fecha) == hoy).all()
    total_hoy  = sum(v.total_venta for v in ventas_hoy)
    top = (db.session.query(Venta.repuesto_id, func.sum(Venta.ganancia_operacion).label('g'))
           .group_by(Venta.repuesto_id).order_by(db.text('g DESC')).first())
    estrella  = db.session.get(Repuesto, top[0]).nombre if top else 'Sin datos'
    todos     = Repuesto.query.all()
    faltantes = [p for p in todos if p.stock < 10]
    inversion = sum(p.p_costo * p.stock for p in todos) if puede_ver_costos() else None
    return dict(
        total_hoy=total_hoy, estrella=estrella, inversion=inversion,
        ganancia=db.session.query(func.sum(Venta.ganancia_operacion)).scalar() or 0,
        inversion_sugerida=sum((10-p.stock)*p.p_costo for p in faltantes) if puede_ver_costos() else None,
        alertas=[p for p in todos if p.stock <= p.stock_minimo],
        puede_costos=puede_ver_costos(),
    )

@inventario_bp.route('/')
def inicio():
    busqueda   = request.args.get('buscar','').upper().strip()
    inventario = (Repuesto.query.filter(func.upper(Repuesto.nombre).contains(busqueda)).all()
                  if busqueda else Repuesto.query.all())
    ventas = Venta.query.order_by(Venta.fecha.desc()).limit(15).all()
    return render_template('index.html', inventario=inventario, ventas=ventas, **_stats())

@inventario_bp.route('/agregar', methods=['POST'])
def agregar():
    if not puede_ver_costos():
        flash('⛔ Sin permiso para agregar productos.', 'danger')
        return redirect(url_for('inventario.inicio'))
    try:
        nombre  = request.form['nombre'].upper().strip()
        p_costo = float(request.form['p_costo'])
        p_venta = float(request.form['p_venta'])
        stock   = int(request.form['stock'])
        if p_costo <= 0 or p_venta <= 0 or stock < 0:
            flash('Los valores deben ser positivos.', 'warning')
            return redirect(url_for('inventario.inicio'))

        # Subir foto si viene
        foto_url = None
        file = request.files.get('foto')
        if file and file.filename and allowed_file(file.filename):
            fname = secure_filename(f"{nombre.replace(' ','_')}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.{file.filename.rsplit('.',1)[1].lower()}")
            upload_folder = os.path.join(current_app.static_folder, 'uploads')
            os.makedirs(upload_folder, exist_ok=True)
            file.save(os.path.join(upload_folder, fname))
            foto_url = f'uploads/{fname}'

        p = Repuesto.query.filter_by(nombre=nombre).first()
        if p:
            ant = p.stock; p.stock += stock; p.p_costo = p_costo; p.p_venta = p_venta
            if foto_url: p.foto_url = foto_url
            _historial(p.id, ant, p.stock, 'AGREGADO')
        else:
            p = Repuesto(nombre=nombre, p_costo=p_costo, p_venta=p_venta, stock=stock, foto_url=foto_url)
            db.session.add(p); db.session.flush()
            _historial(p.id, 0, stock, 'AGREGADO')
        db.session.commit()
        flash(f'✅ {nombre} registrado.', 'success')
    except ValueError:
        flash('❌ Formato numérico incorrecto.', 'danger')
    except Exception as e:
        db.session.rollback(); flash(f'❌ Error: {e}', 'danger')
    return redirect(url_for('inventario.inicio'))

@inventario_bp.route('/editar/<int:id>', methods=['POST'])
def editar(id):
    if not puede_ver_costos():
        flash('⛔ Sin permiso.', 'danger')
        return redirect(url_for('inventario.inicio'))
    try:
        p = db.session.get(Repuesto, id)
        if not p: flash('No encontrado.', 'danger'); return redirect(url_for('inventario.inicio'))
        ant = p.stock
        p.nombre  = request.form['nombre'].upper().strip()
        p.p_costo = float(request.form['p_costo'])
        p.p_venta = float(request.form['p_venta'])
        p.stock   = int(request.form['stock'])

        file = request.files.get('foto')
        if file and file.filename and allowed_file(file.filename):
            fname = secure_filename(f"{p.nombre.replace(' ','_')}_{datetime.utcnow().strftime('%Y%m%d%H%M%S')}.{file.filename.rsplit('.',1)[1].lower()}")
            upload_folder = os.path.join(current_app.static_folder, 'uploads')
            os.makedirs(upload_folder, exist_ok=True)
            file.save(os.path.join(upload_folder, fname))
            p.foto_url = f'uploads/{fname}'

        _historial(p.id, ant, p.stock, 'EDITADO')
        db.session.commit(); flash(f'✅ {p.nombre} actualizado.', 'success')
    except ValueError:
        flash('❌ Formato incorrecto.', 'danger')
    except Exception as e:
        db.session.rollback(); flash(f'❌ {e}', 'danger')
    return redirect(url_for('inventario.inicio'))

@inventario_bp.route('/eliminar_producto/<int:id>')
def eliminar_producto(id):
    if not puede_ver_costos():
        flash('⛔ Sin permiso.', 'danger')
        return redirect(url_for('inventario.inicio'))
    try:
        p = db.session.get(Repuesto, id)
        if not p: flash('No encontrado.', 'danger'); return redirect(url_for('inventario.inicio'))
        db.session.delete(p); db.session.commit()
        flash(f'🗑️ {p.nombre} eliminado.', 'success')
    except Exception as e:
        db.session.rollback(); flash(f'❌ {e}', 'danger')
    return redirect(url_for('inventario.inicio'))

# ── Eliminar masivo ────────────────────────────────────────
@inventario_bp.route('/eliminar_masivo', methods=['POST'])
def eliminar_masivo():
    if not puede_ver_costos():
        return jsonify({'ok': False, 'msg': '⛔ Sin permiso.'})
    try:
        ids = request.json.get('ids', [])
        if not ids:
            return jsonify({'ok': False, 'msg': 'No se seleccionó ningún producto.'})
        eliminados = 0
        for id in ids:
            p = db.session.get(Repuesto, id)
            if p:
                db.session.delete(p)
                eliminados += 1
        db.session.commit()
        return jsonify({'ok': True, 'msg': f'🗑️ {eliminados} productos eliminados.'})
    except Exception as e:
        db.session.rollback()
        return jsonify({'ok': False, 'msg': str(e)})

@inventario_bp.route('/subir_masivo', methods=['POST'])
def subir_masivo():
    if not puede_ver_costos():
        flash('⛔ Sin permiso.', 'danger')
        return redirect(url_for('inventario.inicio'))
    file = request.files.get('archivo_excel')
    if not file: flash('No se seleccionó archivo.', 'warning'); return redirect(url_for('inventario.inicio'))
    try:
        df = pd.read_excel(file)
        for _, row in df.iterrows():
            nom = str(row['nombre']).upper().strip()
            p   = Repuesto.query.filter_by(nombre=nom).first()
            if p:
                ant = p.stock; p.stock += int(row['stock'])
                p.p_costo = float(row['p_costo']); p.p_venta = float(row['p_venta'])
                _historial(p.id, ant, p.stock, 'AGREGADO')
            else:
                p = Repuesto(nombre=nom, p_costo=float(row['p_costo']),
                             p_venta=float(row['p_venta']), stock=int(row['stock']))
                db.session.add(p); db.session.flush()
                _historial(p.id, 0, int(row['stock']), 'AGREGADO')
        db.session.commit(); flash('📁 Excel cargado.', 'success')
    except Exception as e:
        db.session.rollback(); flash(f'❌ Error: {e}', 'danger')
    return redirect(url_for('inventario.inicio'))

@inventario_bp.route('/exportar_compras')
def exportar_compras():
    faltantes = Repuesto.query.filter(Repuesto.stock < 10).all()
    if not faltantes: return '✅ Inventario al día.'
    return '\n'.join(['🛒 PEDIDO DRIVEFLOW PRO'] + [f'• {p.nombre}: Pedir {10-p.stock} uds' for p in faltantes])

@inventario_bp.route('/exportar_inventario')
def exportar_inventario():
    if not puede_ver_costos():
        flash('⛔ Sin permiso para exportar.', 'danger')
        return redirect(url_for('inventario.inicio'))
    try:
        import io
        from flask import send_file
        from datetime import datetime

        productos = Repuesto.query.order_by(Repuesto.nombre).all()

        data = [{
            'Nombre':        p.nombre,
            'Costo ($)':     round(p.p_costo, 2),
            'Venta ($)':     round(p.p_venta, 2),
            'Margen (%)':    round((p.p_venta - p.p_costo) / p.p_costo * 100, 1) if p.p_costo > 0 else 0,
            'Stock':         p.stock,
            'Stock Mínimo':  p.stock_minimo,
            'Vendido':       p.vendido,
            'Estado':        p.estado_stock.upper(),
            'Valor en stock ($)': round(p.p_costo * p.stock, 2),
        } for p in productos]

        df = pd.DataFrame(data)

        # Totales al final
        totales = pd.DataFrame([{
            'Nombre':        'TOTAL',
            'Costo ($)':     '',
            'Venta ($)':     '',
            'Margen (%)':    '',
            'Stock':         df['Stock'].sum(),
            'Stock Mínimo':  '',
            'Vendido':       df['Vendido'].sum(),
            'Estado':        '',
            'Valor en stock ($)': round(df['Valor en stock ($)'].sum(), 2),
        }])
        df = pd.concat([df, totales], ignore_index=True)

        buf = io.BytesIO()
        with pd.ExcelWriter(buf, engine='openpyxl') as writer:
            df.to_excel(writer, index=False, sheet_name='Inventario')

            # Dar formato a la hoja
            ws = writer.sheets['Inventario']
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter

            # Cabeceras
            header_fill = PatternFill('solid', fgColor='0F172A')
            header_font = Font(bold=True, color='38BDF8', size=11)
            for cell in ws[1]:
                cell.fill      = header_fill
                cell.font      = header_font
                cell.alignment = Alignment(horizontal='center')

            # Fila de totales
            last_row = ws.max_row
            total_fill = PatternFill('solid', fgColor='1E293B')
            total_font = Font(bold=True, color='10B981', size=11)
            for cell in ws[last_row]:
                cell.fill = total_fill
                cell.font = total_font

            # Ancho de columnas
            for col in ws.columns:
                max_len = max(len(str(c.value or '')) for c in col) + 4
                ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len, 30)

            # Colorear filas según estado
            danger_fill  = PatternFill('solid', fgColor='FEE2E2')
            warning_fill = PatternFill('solid', fgColor='FEF3C7')
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row - 1):
                estado = str(row[7].value or '')
                if 'AGOTADO' in estado or 'CRITICO' in estado:
                    for cell in row: cell.fill = danger_fill
                elif 'BAJO' in estado:
                    for cell in row: cell.fill = warning_fill

        buf.seek(0)
        fname = f'inventario_driveflow_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.xlsx'
        return send_file(buf, as_attachment=True, download_name=fname,
                         mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

    except Exception as e:
        flash(f'❌ Error al exportar: {e}', 'danger')
        return redirect(url_for('inventario.inicio'))

@inventario_bp.route('/inventario_pdf')
def inventario_pdf():
    if not puede_ver_costos():
        flash('⛔ Sin permiso.', 'danger')
        return redirect(url_for('inventario.inicio'))
    try:
        import io
        from flask import send_file
        from reportlab.lib.pagesizes import A4, landscape
        from reportlab.lib import colors
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import cm
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

        productos = Repuesto.query.order_by(Repuesto.nombre).all()
        buf = io.BytesIO()
        doc = SimpleDocTemplate(buf, pagesize=landscape(A4),
                                rightMargin=1.5*cm, leftMargin=1.5*cm,
                                topMargin=1.5*cm, bottomMargin=1.5*cm)

        c_dark   = colors.HexColor('#0f172a')
        c_blue   = colors.HexColor('#38bdf8')
        c_green  = colors.HexColor('#10b981')
        c_yellow = colors.HexColor('#fef3c7')
        c_red    = colors.HexColor('#fee2e2')
        c_light  = colors.HexColor('#f8fafc')
        c_gray   = colors.HexColor('#64748b')
        c_border = colors.HexColor('#e2e8f0')

        styles   = getSampleStyleSheet()
        story    = []

        # Título
        story.append(Paragraph(
            '<b>DRIVEFLOW PRO — Inventario de Repuestos</b>',
            ParagraphStyle('t', parent=styles['Normal'], fontSize=16, textColor=c_dark, spaceAfter=4)
        ))
        story.append(Paragraph(
            f'Generado el {datetime.utcnow().strftime("%d/%m/%Y %H:%M")} · {len(productos)} productos',
            ParagraphStyle('s', parent=styles['Normal'], fontSize=9, textColor=c_gray, spaceAfter=14)
        ))

        # Stats resumen
        total_productos = len(productos)
        total_stock     = sum(p.stock for p in productos)
        valor_inventario = sum(p.p_costo * p.stock for p in productos)
        agotados        = sum(1 for p in productos if p.stock == 0)
        criticos        = sum(1 for p in productos if 0 < p.stock <= p.stock_minimo)

        stats = Table([[
            f'Total productos: {total_productos}',
            f'Total en stock: {total_stock} uds',
            f'Valor inventario: ${valor_inventario:,.2f}',
            f'Agotados: {agotados}',
            f'Stock crítico: {criticos}',
        ]], colWidths=[5.5*cm]*5)
        stats.setStyle(TableStyle([
            ('BACKGROUND',   (0,0), (-1,-1), c_dark),
            ('TEXTCOLOR',    (0,0), (-1,-1), c_light),
            ('FONTSIZE',     (0,0), (-1,-1), 8),
            ('FONTNAME',     (0,0), (-1,-1), 'Helvetica-Bold'),
            ('ALIGN',        (0,0), (-1,-1), 'CENTER'),
            ('TOPPADDING',   (0,0), (-1,-1), 8),
            ('BOTTOMPADDING',(0,0), (-1,-1), 8),
            ('ROUNDEDCORNERS', [4]),
        ]))
        story.append(stats)
        story.append(Spacer(1, 0.4*cm))

        # Tabla principal
        headers = ['#', 'Nombre', 'Costo $', 'Venta $', 'Margen %', 'Stock', 'Mínimo', 'Vendido', 'Estado', 'Valor stock $']
        rows = [headers]
        for i, p in enumerate(productos, 1):
            margen = round((p.p_venta - p.p_costo) / p.p_costo * 100, 1) if p.p_costo > 0 else 0
            estado = p.estado_stock.upper()
            rows.append([
                str(i),
                p.nombre[:35],
                f'${p.p_costo:,.2f}',
                f'${p.p_venta:,.2f}',
                f'{margen}%',
                str(p.stock),
                str(p.stock_minimo),
                str(p.vendido),
                estado,
                f'${p.p_costo * p.stock:,.2f}',
            ])

        # Fila totales
        rows.append([
            '', 'TOTAL', '', '', '',
            str(total_stock), '', str(sum(p.vendido for p in productos)),
            '', f'${valor_inventario:,.2f}'
        ])

        col_w = [0.7*cm, 7*cm, 2*cm, 2*cm, 2*cm, 1.5*cm, 1.5*cm, 1.8*cm, 2*cm, 2.5*cm]
        tabla = Table(rows, colWidths=col_w, repeatRows=1)

        # Estilos base
        tabla_style = [
            ('BACKGROUND',    (0,0), (-1,0), c_dark),
            ('TEXTCOLOR',     (0,0), (-1,0), c_light),
            ('FONTNAME',      (0,0), (-1,0), 'Helvetica-Bold'),
            ('FONTSIZE',      (0,0), (-1,-1), 8),
            ('ALIGN',         (2,0), (-1,-1), 'RIGHT'),
            ('ALIGN',         (0,0), (1,-1), 'LEFT'),
            ('ALIGN',         (4,0), (8,-1), 'CENTER'),
            ('ROWBACKGROUNDS',(0,1), (-1,-2), [colors.white, c_light]),
            ('GRID',          (0,0), (-1,-1), 0.3, c_border),
            ('TOPPADDING',    (0,0), (-1,-1), 5),
            ('BOTTOMPADDING', (0,0), (-1,-1), 5),
            ('LEFTPADDING',   (0,0), (-1,-1), 5),
            # Fila totales
            ('BACKGROUND',    (0,-1), (-1,-1), c_dark),
            ('TEXTCOLOR',     (0,-1), (-1,-1), c_light),
            ('FONTNAME',      (0,-1), (-1,-1), 'Helvetica-Bold'),
        ]

        # Colorear filas por estado
        for i, p in enumerate(productos, 1):
            if p.stock == 0:
                tabla_style.append(('BACKGROUND', (0,i), (-1,i), c_red))
            elif p.stock <= p.stock_minimo:
                tabla_style.append(('BACKGROUND', (0,i), (-1,i), c_red))
            elif p.stock <= p.stock_minimo * 2:
                tabla_style.append(('BACKGROUND', (0,i), (-1,i), c_yellow))

        tabla.setStyle(TableStyle(tabla_style))
        story.append(tabla)

        # Pie
        story.append(Spacer(1, 0.5*cm))
        story.append(Paragraph(
            f'<font color="#94a3b8">DriveFlow PRO · Reporte de inventario · {datetime.utcnow().strftime("%d/%m/%Y %H:%M")}</font>',
            ParagraphStyle('pie', parent=styles['Normal'], fontSize=7, alignment=1)
        ))

        doc.build(story)
        buf.seek(0)
        fname = f'inventario_{datetime.utcnow().strftime("%Y%m%d_%H%M")}.pdf'
        return send_file(buf, as_attachment=True, download_name=fname, mimetype='application/pdf')

    except ImportError:
        return '❌ Instala reportlab: pip install reportlab', 500
    except Exception as e:
        flash(f'❌ Error al generar PDF: {e}', 'danger')
        return redirect(url_for('inventario.inicio'))